import os
from openpyxl import load_workbook
import logging
from dotenv import load_dotenv
from util.atomic_excel_operations import atomic_excel_operation

load_dotenv()

MAIN_LEDGER_FILE = os.getenv('MAIN_LEDGER_FILEPATH')

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def perform_main_ledger_update(workbook, employee_name: str, employee_accountNo: str, institution_name: str, date: str, ledger_interest_column: str, ledger_debit_column: str, capital: float = None, interest: float = None):
    logger.info(f"Starting main ledger update for employee: {employee_name}, institution: {institution_name}")
    logger.info(f"Parameters - capital: {capital}, interest: {interest}, date: {date}")
    
    if (capital is None or capital == 0) and (interest is None or interest == 0):
        logger.info(f"No capital or interest amount provided for {employee_name}, skipping main ledger update")
        return {"success": True, "message": "No amounts to update", "action": "skipped"}
    
    logger.info("Getting active worksheet from workbook")
    ws = workbook.active
    logger.info(f"Worksheet max_row: {ws.max_row}")
    
    logger.info("Converting column letters to column numbers")
    try:
        if ledger_interest_column:
            from openpyxl.utils import column_index_from_string
            interest_col_num = column_index_from_string(ledger_interest_column)
            logger.info(f"Interest column '{ledger_interest_column}' converted to column number: {interest_col_num}")
        else:
            raise ValueError("ledger interest column variable not set")
            
        if ledger_debit_column:
            debit_col_num = column_index_from_string(ledger_debit_column)
            logger.info(f"Debit column '{ledger_debit_column}' converted to column number: {debit_col_num}")
        else:
            raise ValueError("ledger debit column variable not set")
    except Exception as e:
        logger.error(f"Error converting column letters to numbers: {str(e)}")
        raise ValueError(f"Error converting column letters to numbers: {str(e)}")
    
    institution_row = None
    logger.info(f"Starting search for institution '{institution_name}' in column L (column 12)")
    
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=12).value
        logger.debug(f"Row {row}, Column L value: '{cell_value}'")
        
        if cell_value:
            cell_value_clean = str(cell_value).strip()
            institution_name_clean = institution_name.strip()
            logger.debug(f"Comparing: '{cell_value_clean.lower()}' with '{institution_name_clean.lower()}'")
            
            if cell_value_clean.lower() == institution_name_clean.lower():
                institution_row = row
                logger.info(f"FOUND institution '{institution_name}' at row {row}")
                break
        
        if row % 50 == 0:
            logger.debug(f"Searched {row} rows for institution, continuing...")
    
    if institution_row is None:
        logger.error(f"Institution '{institution_name}' NOT FOUND in column L after searching {ws.max_row} rows")
        raise ValueError(f"Institution '{institution_name}' not found in column L")
    
    employee_row = None
    empty_count = 0
    search_start_row = institution_row + 1
    search_end_row = ws.max_row + 10
    logger.info(f"Starting search for employee '{employee_name}' from row {search_start_row} to {search_end_row}")
    
    for row in range(search_start_row, search_end_row):
        cell_value = ws.cell(row=row, column=12).value
        cell_value_accountNo = ws.cell(row=row, column=18).value  
        logger.debug(f"Row {row}, Column L value: '{cell_value}'")
        
        if cell_value in (None, ""):
            empty_count += 1
            logger.debug(f"Empty cell found at row {row}, empty count: {empty_count}")
            if empty_count >= 5:
                logger.info(f"Found 5 consecutive empty cells, terminating search at row {row}")
                break
        else:
            empty_count = 0
            cell_value_str = str(cell_value).strip() 
            cell_value_accountNo_str = str(cell_value_accountNo).strip() 
            logger.debug(f"Non-empty cell at row {row}: '{cell_value_str}'")
            
            # if not any(char.islower() for char in cell_value_str.replace(' ', '')):
            #     logger.info(f"Found what appears to be another institution '{cell_value_str}' at row {row}, stopping search")
            #     break
            
            employee_name_clean = employee_name.strip()
            employee_accountNo_clean = employee_accountNo.strip()   
            logger.debug(f"Comparing employee name: '{cell_value_str.lower()}' with '{employee_name_clean.lower()}'")
            logger.debug(f"Comparing employee account number: '{cell_value_accountNo_str.lower()}' with '{employee_accountNo_clean.lower()}'")
            
            if cell_value_str.lower() == employee_name_clean.lower() and cell_value_accountNo_str.lower() == employee_accountNo_clean.lower():
                employee_row = row
                logger.info(f"FOUND employee '{employee_name}' with employee account number '{employee_accountNo}' at row {row}")
                break
    
    if employee_row is None:
        logger.error(f"Employee '{employee_name}' NOT FOUND under institution '{institution_name}' in column L")
        raise ValueError(f"Employee '{employee_name}' not found under institution '{institution_name}' in column L")
    
    updates_made = []
    logger.info("Starting to update interest and capital amounts")
    
    if interest is not None and interest != 0:
        logger.info(f"Updating interest amount: {interest}")
        interest_cell = ws.cell(row=employee_row, column=interest_col_num)
        current_interest = interest_cell.value
        logger.info(f"Current interest value in cell: '{current_interest}'")
        
        if current_interest is None or current_interest == "":
            current_interest = 0.0
            logger.info("Current interest was None/empty, treating as 0.0")
        else:
            try:
                current_interest = float(current_interest)
                logger.info(f"Successfully converted current interest to float: {current_interest}")
            except (ValueError, TypeError):
                logger.warning(f"Invalid interest value '{current_interest}' in cell, treating as 0")
                current_interest = 0.0
        
        new_interest = current_interest + interest
        interest_cell.value = new_interest
        
        logger.info(f"Updated interest for {employee_name}: {current_interest} + {interest} = {new_interest} (Column {ledger_interest_column}, Row {employee_row})")
        updates_made.append(f"interest: {current_interest} + {interest} = {new_interest}")
    
    if capital is not None and capital != 0:
        logger.info(f"Updating capital amount: {capital}")
        debit_cell = ws.cell(row=employee_row, column=debit_col_num)
        current_debit = debit_cell.value
        logger.info(f"Current debit value in cell: '{current_debit}'")
        
        if current_debit is None or current_debit == "":
            current_debit = 0.0
            logger.info("Current debit was None/empty, treating as 0.0")
        else:
            try:
                current_debit = float(current_debit)
                logger.info(f"Successfully converted current debit to float: {current_debit}")
            except (ValueError, TypeError):
                logger.warning(f"Invalid debit value '{current_debit}' in cell, treating as 0")
                current_debit = 0.0
        
        new_debit = current_debit + capital
        debit_cell.value = new_debit
        
        logger.info(f"Updated capital for {employee_name}: {current_debit} + {capital} = {new_debit} (Column {ledger_debit_column}, Row {employee_row})")
        updates_made.append(f"capital: {current_debit} + {capital} = {new_debit}")
    
    logger.info(f"Main ledger update completed successfully for {employee_name}")
    return {
        "success": True,
        "message": f"Main ledger updated for {employee_name}",
        "action": "updated",
        "row_updated": employee_row,
        "updates_made": updates_made,
        "institution_row": institution_row
    }


def update_main_ledger(employee_name: str, employee_accountNo: str, institution_name: str, date: str, ledger_debit_column: str , ledger_interest_column: str ,capital: float = None, interest: float = None) -> dict:
    logger.info("=== STARTING MAIN LEDGER UPDATE ===")
    logger.info(f"Employee: {employee_name}")
    logger.info(f"Institution: {institution_name}")
    logger.info(f"Account No: {employee_accountNo}")
    logger.info(f"Date: {date}")
    logger.info(f"Ledger Debit Column: {ledger_debit_column}")
    logger.info(f"Ledger Interest Column: {ledger_interest_column}")
    logger.info(f"Capital: {capital}")
    logger.info(f"Interest: {interest}")
    
    try:
        logger.info("Validating environment variables...")
        if not MAIN_LEDGER_FILE:
            logger.error("MAIN_LEDGER_FILEPATH environment variable not set")
            raise ValueError("MAIN_LEDGER_FILEPATH environment variable not set")
        else:
            logger.info(f"Main ledger file path: {MAIN_LEDGER_FILE}")
        

        if not ledger_interest_column:
            logger.error("Ledger interest column not provided")
            raise ValueError("Ledger interest column not provided")
        else:
            logger.info(f"Interest column: {ledger_interest_column}")
        
        if not ledger_debit_column:
            logger.error("Ledger debit column not provided")
            raise ValueError("Ledger debit column not provided")
        else:
            logger.info(f"Debit column: {ledger_debit_column}")
        
        logger.info("Checking if main ledger file exists...")
        if not os.path.exists(MAIN_LEDGER_FILE):
            logger.error(f"Main ledger file not found: {MAIN_LEDGER_FILE}")
            raise FileNotFoundError(f"Main ledger file not found: {MAIN_LEDGER_FILE}")
        else:
            logger.info(f"Main ledger file exists: {MAIN_LEDGER_FILE}")
        
        logger.info("Starting atomic Excel operation for main ledger update")
        
        with atomic_excel_operation(MAIN_LEDGER_FILE) as workbook:
            logger.info("Successfully opened workbook with atomic operation")
            result = perform_main_ledger_update(
                workbook, 
                employee_name, 
                employee_accountNo, 
                institution_name, 
                date, 
                ledger_interest_column, 
                ledger_debit_column,
                capital, 
                interest
            )
            logger.info("Main ledger update operation completed")
        
        success_message = f"Successfully updated main ledger for {employee_name}"
        logger.info(success_message)
        logger.info("=== MAIN LEDGER UPDATE COMPLETED SUCCESSFULLY ===")
        
        return {
            "success": True,
            "message": success_message,
            "details": result
        }
        
    except ValueError as ve:
        error_message = str(ve)
        logger.error(f"Validation error updating main ledger for {employee_name}: {error_message}")
        logger.error("=== MAIN LEDGER UPDATE FAILED (VALIDATION) ===")
        return {
            "success": False,
            "error": error_message
        }
        
    except FileNotFoundError as fe:
        error_message = f"File not found: {str(fe)}"
        logger.error(f"File error updating main ledger for {employee_name}: {error_message}")
        logger.error("=== MAIN LEDGER UPDATE FAILED (FILE NOT FOUND) ===")
        return {
            "success": False,
            "error": error_message
        }
        
    except Exception as e:
        error_message = f"Error updating main ledger for {employee_name}: {str(e)}"
        logger.error(error_message)
        logger.error("=== MAIN LEDGER UPDATE FAILED (GENERAL ERROR) ===")
        import traceback
        logger.error(traceback.format_exc())
        return {
            "success": False,
            "error": error_message
        }