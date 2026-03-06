import logging
from openpyxl import load_workbook
from util.finding_files_sheets import find_personal_account_file, find_employee_sheet  # Import file and sheet finding functions
import win32com.client
import pythoncom

def force_excel_recalculation(file_path):
    """
    Opens the Excel file using Microsoft Excel,
    forces recalculation of formulas, and saves it.
    """
    # Initialize COM for the current thread
    pythoncom.CoInitialize()
    
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(file_path)

        # Force full recalculation
        excel.CalculateFull()

        wb.Save()
        wb.Close()
        excel.Quit()
    finally:
        # Always uninitialize COM when done to free up resources
        pythoncom.CoUninitialize()

logger = logging.getLogger(__name__)

def validate_capital_limit_xlsx(employee_name: str, institution_name: str, acc_no: str, capital: float):
    """
    Validates if the capital amount exceeds the limit in Column K.
    Finds the file and the next empty row (target row) to check the specific limit for that entry.
    """
    
    # 1. Early exit if no capital to validate
    if capital is None or float(capital) <= 0:
        return

    logger.info(f"Validating capital limit for {employee_name} ({acc_no}) - Amount: {capital}")

    # 2. Find the file path (Reuse existing logic)
    try:
        file_path = find_personal_account_file(employee_name, acc_no, institution_name)
    except FileNotFoundError as e:
        raise e
    
    force_excel_recalculation(file_path)

    # 3. Load Workbook in READ-ONLY and DATA-ONLY mode
    wb = load_workbook(file_path, data_only=True, read_only=True)
    
    try:
        # 4. Find the correct sheet (Reuse existing logic)
        ws = find_employee_sheet(wb, acc_no)
        
        # 5. Find the Target Row (The same logic as perform_personal_account_update)
        current_row = None
        empty_rows_count = 0
        first_empty_row = None
        
        # We iterate to find the 4 consecutive empty rows
        # for row in range(1, ws.max_row + 100): 
        #     # In read_only mode, use ws.cell(row, col).value
            
        #     capital_val = ws.cell(row=row, column=9).value
            
        #     is_row_empty = all(val in (None, "") for val in [capital_val])
            
        #     if is_row_empty:
        #         if empty_rows_count == 0:
        #             first_empty_row = row
        #         empty_rows_count += 1
                
        #         if empty_rows_count >= 4:
        #             current_row = first_empty_row
        #             break
        #     else:
        #         empty_rows_count = 0
        #         first_empty_row = None

        for row in range(1, ws.max_row + 100):

            capital_val = ws.cell(row=row, column=9).value

            is_row_empty = capital_val is None or str(capital_val).strip() == ""

            if is_row_empty:
                if empty_rows_count == 0:
                    first_empty_row = row

                empty_rows_count += 1

                if empty_rows_count >= 4:
                    current_row = first_empty_row
                    break

            else:
                empty_rows_count = 0
                first_empty_row = None

        if current_row is None:
            raise ValueError(f"Could not find available rows to validate limit for {employee_name}")

        # 6. Read the Limit from Column K (Column 11) of the target row
        limit_val = ws.cell(row=current_row, column=11).value

        if limit_val is None:
            limit_val = ws.cell(row=current_row - 1, column=11).value

        if limit_val is None:
            limit_val = ws.cell(row=current_row - 2, column=11).value


        logger.info("The limit_val read from the sheet is: %s", limit_val)
        
        # Handle conversion safely
        try:
            limit_float = float(limit_val) if limit_val is not None else 0.0
        except (ValueError, TypeError):
            # If the formula evaluates to error or string, treat limit as 0 or handle accordingly
            limit_float = 0.0

        logger.info(f"Row {current_row} Limit: {limit_float}, Requested Capital: {capital}")

        # 7. Compare
        if limit_float < float(capital):
            raise ValueError(
                f"Capital limit reached! Limit is {limit_float}, but attempted to pay {capital}."
            )

    finally:
        # Explicitly close the read-only workbook to free memory/file handles
        wb.close()