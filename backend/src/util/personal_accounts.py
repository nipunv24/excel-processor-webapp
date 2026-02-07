import os
from openpyxl import load_workbook
import xlrd
import xlwt
from xlutils.copy import copy
import logging
from dotenv import load_dotenv
from util.atomic_excel_operations import atomic_excel_operation  # Import our atomic operations
from util.validate_capital_limit_utilities import validate_capital_limit_xlsx  # Import the capital limit validation function
from util.finding_files_sheets import find_personal_account_file, find_employee_sheet_xls, find_employee_sheet  # Import file and sheet finding functions


load_dotenv()

PERSONAL_ACCOUNT_ROOTPATH = os.getenv('PERSONAL_ACCOUNT_ROOTPATH')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)





def perform_personal_account_update_xls(file_path: str, employee_name: str, employee_accountNo: str, date: str, capital: float = None, interest: float = None, description: str = None, bill_no: str = "BS", cheque_no: str = "") -> int:
    """
    Handle .xls files using xlrd/xlwt
    
    Args:
        file_path (str): Full path to the .xls file
        employee_name (str): Name of the employee
        employee_accountNo (str): Account number of the employee
        date (str): Date of the payment
        capital (float, optional): Capital amount. Defaults to None.
        interest (float, optional): Interest amount. Defaults to None.
        description (str, optional): Description for the entry. Defaults to None.
        
    Returns:
        int: The row number that was updated
    """
    
    # Read the existing file
    rb = xlrd.open_workbook(file_path, formatting_info=True)
    
    # Find the correct sheet for this employee
    sheet_index, sheet = find_employee_sheet_xls(rb, employee_accountNo)
    
    # Find 4 consecutive empty rows
    current_row = None
    empty_rows_count = 0
    first_empty_row = None
    
    max_rows = max(sheet.nrows + 100, 1000)  # Ensure we check enough rows
    
    for row in range(max_rows):
        # Check if current row is empty in columns A (0), H (7), and I (8)
        date_value = ""
        interest_value = ""
        capital_value = ""
        
        if row < sheet.nrows:
            if sheet.ncols > 0:
                date_value = sheet.cell_value(row, 0)
            if sheet.ncols > 7:
                interest_value = sheet.cell_value(row, 7)
            if sheet.ncols > 8:
                capital_value = sheet.cell_value(row, 8)
        
        is_row_empty = all(
            str(value).strip() == "" 
            for value in [date_value, interest_value, capital_value]
        )
        
        if is_row_empty:
            if empty_rows_count == 0:
                first_empty_row = row
            empty_rows_count += 1
            
            if empty_rows_count >= 4:
                current_row = first_empty_row
                break
        else:
            empty_rows_count = 0
            first_empty_row = None
    
    if current_row is None:
        raise ValueError(f"Could not find 4 consecutive empty rows in personal account file for {employee_name}")
    
    # Create a copy of the workbook for writing
    wb = copy(rb)
    ws = wb.get_sheet(sheet_index)  # Use the found sheet index
    
    # Update the cells
    ws.write(current_row, 0, date)  # Date in Column A (0)
    ws.write(current_row, 1, bill_no)   # Column B (1) - Bill No
    ws.write(current_row, 2, cheque_no) # Column C (2) - Cheque No
    
    if interest is not None:
        ws.write(current_row, 7, interest)  # Interest in Column H (7)
        
    if capital is not None:
        ws.write(current_row, 8, capital)  # Capital in Column I (8)

    if description is not None:
        ws.write(current_row, 4, description)  # Description in Column E (4)
    
    # Save the file
    wb.save(file_path)
    
    return current_row





def perform_personal_account_update(workbook, file_path:str, employee_name: str, employee_accountNo: str, institution_name: str, date: str, capital: float = None, interest: float = None, description: str = None, bill_no: str = "BS", cheque_no: str = "") -> int:
    """
    Separated personal account update logic to work with atomic operations
    
    Args:
        workbook: The openpyxl workbook object to work with
        employee_name (str): Name of the employee
        employee_accountNo (str): Account number of the employee
        institution_name (str): Name of the institution
        date (str): Date of the payment
        capital (float, optional): Capital amount. Defaults to None.
        interest (float, optional): Interest amount. Defaults to None.
        description (str, optional): Description for the entry. Defaults to None.
        
    Returns:
        int: The row number that was updated
    """
    
    # Find the correct sheet for this employee
    ws = find_employee_sheet(workbook, employee_accountNo)
    
    # Find 4 consecutive empty rows
    current_row = None
    empty_rows_count = 0
    first_empty_row = None
    
    for row in range(1, ws.max_row + 100):  # +100 to ensure we check enough rows
        # Check if current row is empty in columns A, H, and I
        date_cell = ws.cell(row=row, column=1)  # Column A
        interest_cell = ws.cell(row=row, column=8)  # Column H
        capital_cell = ws.cell(row=row, column=9)  # Column I
        
        is_row_empty = all(
            cell.value in (None, "") 
            for cell in [date_cell, interest_cell, capital_cell]
        )
        
        if is_row_empty:
            if empty_rows_count == 0:
                # Remember the first empty row of the sequence
                first_empty_row = row
            empty_rows_count += 1
            
            if empty_rows_count >= 4:
                current_row = first_empty_row
                break
        else:
            # Reset counter if we find a non-empty row
            empty_rows_count = 0
            first_empty_row = None
    
    if current_row is None:
        raise ValueError(f"Could not find 4 consecutive empty rows in personal account file for {employee_name}")
    
        
    # Update the cells
    # Date in Column A
    ws.cell(row=current_row, column=1).value = date

    # Bill No in Column B (2) - Replaces the hardcoded "BS"
    ws.cell(row=current_row, column=2).value = bill_no
    
    # Cheque No in Column C (3)
    ws.cell(row=current_row, column=3).value = cheque_no
    
    # Interest in Column H (if provided)
    if interest is not None:
        ws.cell(row=current_row, column=8).value = interest
        
    # Capital in Column I (if provided)
    if capital is not None:
        ws.cell(row=current_row, column=9).value = capital
    
    if description is not None:
        ws.cell(row=current_row, column=4).value = description
    # Updating column 2 with the word BS
   
    
    return current_row



def update_personal_account(employee_name: str, employee_accountNo: str, institution_name: str, date: str, capital: float = None, interest: float = None, description: str = None,bill_no: str = "BS",cheque_no: str = "") -> dict:
    """
    Updates the personal account Excel file for a specific employee with payment information.
    Looks for 4 consecutive empty rows and uses the first one for the update.
    Uses atomic operations to prevent file corruption and flexible file matching.
    
    Args:
        employee_name (str): Name of the employee
        employee_accountNo (str): Account number of the employee
        institution_name (str): Name of the institution
        date (str): Date of the payment
        capital (float, optional): Capital amount. Defaults to None.
        interest (float, optional): Interest amount. Defaults to None.
        
    Returns:
        dict: A dictionary containing success status and additional information
    """

    try:
        # Find the file (supports both .xls and .xlsx)

        logger.info(f"********************************************************Personal Account Update Request - Employee: {employee_name}, Bill No: {bill_no}, Cheque No: {cheque_no}********************************************")

        file_path = find_personal_account_file(employee_name, employee_accountNo, institution_name)
        
        logger.info(f"The file path of the employee is {file_path}")
        
        # Determine file type and use appropriate handler
        if file_path.lower().endswith('.xlsx'):
            logger.info("Processing .xlsx file with openpyxl")
            # Use existing atomic operations for .xlsx files
            with atomic_excel_operation(file_path) as workbook:
                current_row = perform_personal_account_update(
                    workbook=workbook, 
                    file_path=file_path,  
                    employee_name=employee_name, 
                    employee_accountNo=employee_accountNo, 
                    institution_name=institution_name, 
                    date=date, 
                    capital=capital, 
                    interest=interest,
                    description=description,
                    bill_no=bill_no,
                    cheque_no=cheque_no
                )
        elif file_path.lower().endswith('.xls'):
            logger.info("Processing .xls file with xlrd/xlwt")
            current_row = perform_personal_account_update_xls(
                file_path, 
                employee_name, 
                date, 
                capital, 
                interest,
                description,
                bill_no,
                cheque_no
            )
        else:
            raise ValueError(f"Unsupported file format: {file_path}")
        
        success_message = f"Successfully updated personal account for {employee_name} at row {current_row}"
        logger.info(success_message)
        
        return {
            "success": True,
            "message": success_message,
            "row_updated": current_row,
            "file_path": file_path
        }
        
    except ValueError as ve:
        # Handle specific ValueError (like not finding empty rows)
        error_message = str(ve)
        logger.error(f"Validation error updating personal account for {employee_name}: {error_message}")
        return {
            "success": False,
            "error": error_message
        }
        
    except FileNotFoundError as fe:
        # Handle file not found errors
        error_message = f"File not found: {str(fe)}"
        logger.error(f"File error updating personal account for {employee_name}: {error_message}")
        return {
            "success": False,
            "error": error_message
        }
        
    except Exception as e:
        # Handle any other unexpected errors
        error_message = f"Error updating personal account for {employee_name}: {str(e)}"
        logger.error(error_message)
        import traceback
        logger.error(traceback.format_exc())
        return {
            "success": False,
            "error": error_message
        }