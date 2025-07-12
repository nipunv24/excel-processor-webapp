import os
from openpyxl import load_workbook
import xlrd
import xlwt
from xlutils.copy import copy
import logging
from dotenv import load_dotenv
from util.atomic_excel_operations import atomic_excel_operation  # Import our atomic operations

load_dotenv()

TRIAL_BALANCE_FILE = os.getenv('TRIAL_BALANCE_ROOTPATH')
CAPITAL_WORKSHEET = os.getenv('TRIAL_BALANCE_CAPITAL_UPDATE_WORKSHEET_NAME')
INTEREST_WORKSHEET = os.getenv('TRIAL_BALANCE_INTEREST_UPDATE_WORKSHEET_NAME')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def perform_interest_trial_balance_update(workbook, employee_name: str, employee_accountNo: str, institution_name: str, date: str, capital: float = None, interest: float = None):
    """
    Update interest trial balance worksheet logic
    
    Args:
        workbook: The openpyxl workbook object to work with
        employee_name (str): Name of the employee
        employee_accountNo (str): Account number of the employee
        institution_name (str): Name of the institution
        date (str): Date of the payment
        capital (float, optional): Capital amount. Defaults to None.
        interest (float, optional): Interest amount. Defaults to None.
        
    Returns:
        dict: Result information
    """
    
    if interest is None or interest == 0:
        logger.info(f"No interest amount provided for {employee_name}, skipping interest trial balance update")
        return {"success": True, "message": "No interest to update", "action": "skipped"}
    
    # Access the interest worksheet
    if INTEREST_WORKSHEET not in workbook.sheetnames:
        raise ValueError(f"Interest worksheet '{INTEREST_WORKSHEET}' not found in trial balance file")
    
    ws = workbook[INTEREST_WORKSHEET]
    
    # Find 5 consecutive empty rows and select the row before the first empty row
    target_row = None
    empty_rows_count = 0
    first_empty_row = None
    
    for row in range(1, ws.max_row + 10):  # +10 to ensure we check enough rows
        date_cell = ws.cell(row=row, column=1)  # Column A
        
        if date_cell.value in (None, ""):
            if empty_rows_count == 0:
                first_empty_row = row
            empty_rows_count += 1
            
            if empty_rows_count >= 5:
                target_row = first_empty_row
                break
        else:
            empty_rows_count = 0
            first_empty_row = None
    
    if target_row is None:
        raise ValueError("Could not find 5 consecutive empty rows for interest trial balance update")
    
    # Check if there's a previous row to compare dates
    previous_row = target_row - 1
    date_matches = False
    
    if previous_row >= 1:
        previous_date_cell = ws.cell(row=previous_row, column=1)
        if previous_date_cell.value and str(previous_date_cell.value).strip() == str(date).strip():
            date_matches = True
            target_row = previous_row  # Use the existing row
    
    if date_matches:
        # Date matches, add interest to existing value in column F
        interest_cell = ws.cell(row=target_row, column=6)  # Column F
        current_interest = interest_cell.value
        
        # Convert to float, handle None or empty values
        if current_interest is None or current_interest == "":
            current_interest = 0.0
        else:
            current_interest = float(current_interest)
        
        # Add the new interest amount
        new_interest = current_interest + interest
        interest_cell.value = new_interest
        
        logger.info(f"Updated existing interest entry for {date}: {current_interest} + {interest} = {new_interest}")
        action = "updated_existing"
    else:
        # Date doesn't match or no previous row, create new entry
        ws.cell(row=target_row, column=1).value = date  # Column A
        ws.cell(row=target_row, column=6).value = interest  # Column F
        
        logger.info(f"Created new interest entry for {date}: {interest}")
        action = "created_new"
    
    return {
        "success": True,
        "message": f"Interest trial balance updated for {employee_name}",
        "action": action,
        "row_updated": target_row,
        "amount": interest
    }


def perform_capital_trial_balance_update(workbook, employee_name: str, employee_accountNo: str, institution_name: str, date: str, capital: float = None, interest: float = None):
    """
    Update capital trial balance worksheet logic
    
    Args:
        workbook: The openpyxl workbook object to work with
        employee_name (str): Name of the employee
        employee_accountNo (str): Account number of the employee
        institution_name (str): Name of the institution
        date (str): Date of the payment
        capital (float, optional): Capital amount. Defaults to None.
        interest (float, optional): Interest amount. Defaults to None.
        
    Returns:
        dict: Result information
    """
    
    if capital is None or capital == 0:
        logger.info(f"No capital amount provided for {employee_name}, skipping capital trial balance update")
        return {"success": True, "message": "No capital to update", "action": "skipped"}
    
    # Access the capital worksheet
    if CAPITAL_WORKSHEET not in workbook.sheetnames:
        raise ValueError(f"Capital worksheet '{CAPITAL_WORKSHEET}' not found in trial balance file")
    
    logger.info(f"Available worksheets in trial balance file: {workbook.sheetnames}")
    
    ws = workbook[CAPITAL_WORKSHEET]
    
    # Find 5 consecutive empty rows and select the row before the first empty row
    target_row = None
    empty_rows_count = 0
    first_empty_row = None
    
    for row in range(1, ws.max_row + 10):  # +10 to ensure we check enough rows
        date_cell = ws.cell(row=row, column=1)  # Column A
        
        if date_cell.value in (None, ""):
            if empty_rows_count == 0:
                first_empty_row = row
            empty_rows_count += 1
            
            if empty_rows_count >= 5:
                target_row = first_empty_row
                break
        else:
            empty_rows_count = 0
            first_empty_row = None
    
    if target_row is None:
        raise ValueError("Could not find 5 consecutive empty rows for capital trial balance update")
    
    # Check if there's a previous row to compare dates
    previous_row = target_row - 1
    date_matches = False
    
    if previous_row >= 1:
        previous_date_cell = ws.cell(row=previous_row, column=1)
        if previous_date_cell.value and str(previous_date_cell.value).strip() == str(date).strip():
            date_matches = True
            target_row = previous_row  # Use the existing row
    
    if date_matches:
        # Date matches, add capital to existing value in column F
        capital_cell = ws.cell(row=target_row, column=6)  # Column F
        current_capital = capital_cell.value
        
        # Convert to float, handle None or empty values
        if current_capital is None or current_capital == "":
            current_capital = 0.0
        else:
            current_capital = float(current_capital)
        
        # Add the new capital amount
        new_capital = current_capital + capital
        capital_cell.value = new_capital
        
        logger.info(f"Updated existing capital entry for {date}: {current_capital} + {capital} = {new_capital}")
        action = "updated_existing"
    else:
        # Date doesn't match or no previous row, create new entry
        ws.cell(row=target_row, column=1).value = date  # Column A
        ws.cell(row=target_row, column=6).value = capital  # Column F
        
        logger.info(f"Created new capital entry for {date}: {capital}")
        action = "created_new"
    
    return {
        "success": True,
        "message": f"Capital trial balance updated for {employee_name}",
        "action": action,
        "row_updated": target_row,
        "amount": capital
    }


def update_interest_trial_balance(employee_name: str, employee_accountNo: str, institution_name: str, date: str, capital: float = None, interest: float = None) -> dict:
    """
    Updates the interest trial balance Excel file with payment information.
    Uses atomic operations to prevent file corruption.
    
    Args:
        employee_name (str): Name of the employee
        employee_accountNo (str): Account number of the employee
        institution_name (str): Name of the institution
        date (str): Date of the payment
        capital (float, optional): Capital amount. Defaults to None.
        interest (float, optional): Interest amount. Defaults to None.
        
    Returns:
        dict: A dictionary containing success status and additional information
    """
    
    try:
        # Validate that trial balance file exists
        if not os.path.exists(TRIAL_BALANCE_FILE):
            raise FileNotFoundError(f"Trial balance file not found: {TRIAL_BALANCE_FILE}")
        
        logger.info(f"Updating interest trial balance for {employee_name}")
        
        # Use atomic operations for .xlsx files
        with atomic_excel_operation(TRIAL_BALANCE_FILE) as workbook:
            result = perform_interest_trial_balance_update(
                workbook, 
                employee_name, 
                employee_accountNo, 
                institution_name, 
                date, 
                capital, 
                interest
            )
        
        success_message = f"Successfully updated interest trial balance for {employee_name}"
        logger.info(success_message)
        
        return {
            "success": True,
            "message": success_message,
            "details": result
        }
        
    except ValueError as ve:
        error_message = str(ve)
        logger.error(f"Validation error updating interest trial balance for {employee_name}: {error_message}")
        return {
            "success": False,
            "error": error_message
        }
        
    except FileNotFoundError as fe:
        error_message = f"File not found: {str(fe)}"
        logger.error(f"File error updating interest trial balance for {employee_name}: {error_message}")
        return {
            "success": False,
            "error": error_message
        }
        
    except Exception as e:
        error_message = f"Error updating interest trial balance for {employee_name}: {str(e)}"
        logger.error(error_message)
        import traceback
        logger.error(traceback.format_exc())
        return {
            "success": False,
            "error": error_message
        }


def update_capital_trial_balance(employee_name: str, employee_accountNo: str, institution_name: str, date: str, capital: float = None, interest: float = None) -> dict:
    """
    Updates the capital trial balance Excel file with payment information.
    Uses atomic operations to prevent file corruption.
    
    Args:
        employee_name (str): Name of the employee
        employee_accountNo (str): Account number of the employee
        institution_name (str): Name of the institution
        date (str): Date of the payment
        capital (float, optional): Capital amount. Defaults to None.
        interest (float, optional): Interest amount. Defaults to None.
        
    Returns:
        dict: A dictionary containing success status and additional information
    """
    
    try:
        # Validate that trial balance file exists
        if not os.path.exists(TRIAL_BALANCE_FILE):
            raise FileNotFoundError(f"Trial balance file not found: {TRIAL_BALANCE_FILE}")
        
        logger.info(f"Updating capital trial balance for {employee_name}")
        
        # Use atomic operations for .xlsx files
        with atomic_excel_operation(TRIAL_BALANCE_FILE) as workbook:
            result = perform_capital_trial_balance_update(
                workbook, 
                employee_name, 
                employee_accountNo, 
                institution_name, 
                date, 
                capital, 
                interest
            )
        
        success_message = f"Successfully updated capital trial balance for {employee_name}"
        logger.info(success_message)
        
        return {
            "success": True,
            "message": success_message,
            "details": result
        }
        
    except ValueError as ve:
        error_message = str(ve)
        logger.error(f"Validation error updating capital trial balance for {employee_name}: {error_message}")
        return {
            "success": False,
            "error": error_message
        }
        
    except FileNotFoundError as fe:
        error_message = f"File not found: {str(fe)}"
        logger.error(f"File error updating capital trial balance for {employee_name}: {error_message}")
        return {
            "success": False,
            "error": error_message
        }
        
    except Exception as e:
        error_message = f"Error updating capital trial balance for {employee_name}: {str(e)}"
        logger.error(error_message)
        import traceback
        logger.error(traceback.format_exc())
        return {
            "success": False,
            "error": error_message
        }