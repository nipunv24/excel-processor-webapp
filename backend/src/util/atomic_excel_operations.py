import os
import shutil
import tempfile
import logging
from contextlib import contextmanager
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

class AtomicExcelOperation:
    """
    Context manager for atomic Excel file operations.
    Creates a temporary copy, performs operations, and atomically replaces the original.
    """
    
    def __init__(self, original_file_path):
        self.original_file_path = original_file_path
        self.temp_file_path = None
        self.workbook = None
        
    def __enter__(self):
        """
        Create temporary copy and return workbook for operations
        """
        try:
            # Validate that original file exists
            if not os.path.exists(self.original_file_path):
                raise FileNotFoundError(f"Original Excel file not found: {self.original_file_path}")
            
            # Create temporary file with same extension
            file_dir = os.path.dirname(self.original_file_path)
            file_name = os.path.basename(self.original_file_path)
            name, ext = os.path.splitext(file_name)
            
            # Create temporary file in the same directory as original
            temp_fd, self.temp_file_path = tempfile.mkstemp(
                suffix=ext, 
                prefix=f"{name}_temp_", 
                dir=file_dir
            )
            os.close(temp_fd)  # Close the file descriptor
            
            # Copy original file to temporary location
            shutil.copy2(self.original_file_path, self.temp_file_path)
            logger.info(f"Created temporary copy: {self.temp_file_path}")
            
            # Load workbook from temporary file
            self.workbook = load_workbook(self.temp_file_path)
            logger.info("Loaded workbook from temporary file")
            
            return self.workbook
            
        except Exception as e:
            # Cleanup if initialization fails
            self._cleanup_temp_file()
            logger.error(f"Failed to initialize atomic operation: {str(e)}")
            raise
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """
        Save changes and atomically replace original file, or cleanup on error
        """
        try:
            if exc_type is None:
                # No exception occurred, commit the changes
                self._commit_changes()
                logger.info("Atomic operation completed successfully")
            else:
                # Exception occurred, cleanup without committing
                logger.error(f"Exception in atomic operation: {exc_val}")
                self._cleanup_temp_file()
                
        except Exception as commit_error:
            logger.error(f"Error during commit: {str(commit_error)}")
            self._cleanup_temp_file()
            raise commit_error
    
    def _commit_changes(self):
        """
        Save workbook and atomically replace original file
        """
        try:
            if self.workbook:
                # Save changes to temporary file
                self.workbook.save(self.temp_file_path)
                logger.info("Saved changes to temporary file")
                
                # Close workbook to release file handles
                self.workbook.close()
                self.workbook = None
                
                # Create backup of original file (optional safety measure)
                backup_path = f"{self.original_file_path}.backup"
                if os.path.exists(backup_path):
                    os.remove(backup_path)
                shutil.copy2(self.original_file_path, backup_path)
                logger.info(f"Created backup: {backup_path}")
                
                # Atomically replace original file
                if os.name == 'nt':  # Windows
                    # On Windows, we need to remove the target first
                    if os.path.exists(self.original_file_path):
                        os.remove(self.original_file_path)
                    shutil.move(self.temp_file_path, self.original_file_path)
                else:  # Unix/Linux/Mac
                    # On Unix systems, os.rename is atomic
                    os.rename(self.temp_file_path, self.original_file_path)
                
                logger.info("Atomically replaced original file")
                
                # Remove backup after successful operation
                if os.path.exists(backup_path):
                    os.remove(backup_path)
                    
        except Exception as e:
            logger.error(f"Error during commit: {str(e)}")
            self._cleanup_temp_file()
            raise
    
    def _cleanup_temp_file(self):
        """
        Clean up temporary file
        """
        try:
            if self.workbook:
                self.workbook.close()
                self.workbook = None
                
            if self.temp_file_path and os.path.exists(self.temp_file_path):
                os.remove(self.temp_file_path)
                logger.info(f"Cleaned up temporary file: {self.temp_file_path}")
                
        except Exception as e:
            logger.warning(f"Error cleaning up temporary file: {str(e)}")


@contextmanager
def atomic_excel_operation(file_path):
    """
    Convenience function to use atomic Excel operations as a context manager
    
    Usage:
        with atomic_excel_operation(EXCEL_FILE_PATH) as wb:
            ws = wb["Sheet1"]
            ws["A1"] = "New Value"
            # Changes are automatically committed when exiting the context
    """
    atomic_op = AtomicExcelOperation(file_path)
    try:
        wb = atomic_op.__enter__()
        yield wb
    except Exception as e:
        atomic_op.__exit__(type(e), e, e.__traceback__)
        raise
    else:
        atomic_op.__exit__(None, None, None)


def safe_excel_operation(file_path, operation_func, *args, **kwargs):
    """
    Execute an Excel operation safely with atomic file handling
    
    Args:
        file_path: Path to the Excel file
        operation_func: Function that performs Excel operations
        *args, **kwargs: Arguments to pass to operation_func
    
    The operation_func should accept workbook as its first parameter
    """
    with atomic_excel_operation(file_path) as workbook:
        return operation_func(workbook, *args, **kwargs)