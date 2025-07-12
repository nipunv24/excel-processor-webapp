from flask import Flask, request, jsonify
from openpyxl import load_workbook
from flask_cors import CORS
import logging
from util.personal_accounts import update_personal_account
from util.atomic_excel_operations import atomic_excel_operation  # Import our atomic operations
from util.trial_balance_updates import update_capital_trial_balance, update_interest_trial_balance
import os
from dotenv import load_dotenv

load_dotenv()
EXCEL_FILE_PATH = os.getenv('CASHBOOK_FILEPATH')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app, resources={r"/*": {
    "origins": "*", 
    "methods": ["GET", "POST", "OPTIONS"],
    "allow_headers": ["Content-Type", "Authorization"]
}})


@app.route('/update-cell', methods=['POST'])
def update_cell():
    try:
        data = request.json
        sheet_name = data.get("sheet")
        cell = data.get("cell")  # e.g., "B2"
        new_value = data.get("value")

        if not cell or new_value is None:
            return jsonify({"error": "Cell and value are required"}), 400

        # Use atomic operation for Excel file
        with atomic_excel_operation(EXCEL_FILE_PATH) as wb:
            ws = wb[sheet_name]
            ws[cell] = new_value
            # File is automatically saved and atomically replaced when exiting context

        return jsonify({"message": f"Cell {cell} updated successfully!"})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


def perform_payment_operation(workbook, data):
    """
    Separated payment logic to work with atomic operations
    """
    # Extract all form data
    institute = data.get("institute")
    employee = data.get("employee")
    capital_amount = data.get("capitalAmount")
    interest_amount = data.get("interestAmount")
    bill_no = data.get("billNo", "")
    cheq_no = data.get("cheqNo")
    acc_no = data.get("accNo")
    bank_name = data.get("bankName", "")
    description = data.get("description", "")
    first_entry = data.get("firstEntry")
    date = data.get("date")
    
    # Convert first_entry to integer
    fer = int(first_entry)
    
    ws = workbook["Sheet1"]  # Using Sheet1 by default
    
    # Convert amounts to float for numeric handling if they exist
    capital_value = None
    if capital_amount:
        try:
            capital_value = float(capital_amount)
        except ValueError:
            raise ValueError("Capital amount must be a valid number")
            
    interest_value = None
    if interest_amount:
        try:
            interest_value = float(interest_amount)
        except ValueError:
            raise ValueError("Interest amount must be a valid number")
    
    # Check if the cells from Bfer to Jfer are empty
    is_empty_row = True
    for col in range(2, 11):  # B to J columns (2 to 10 in 0-based indexing)
        cell_value = ws.cell(row=fer, column=col).value
        if cell_value not in (None, ""):
            is_empty_row = False
            break
    
    # If the first row isn't empty, find three consecutive empty rows
    current_row = fer
    if not is_empty_row:
        logger.info(f"First entry row {fer} is not empty, searching for three consecutive empty rows...")
        found = False
        
        for row_num in range(fer, ws.max_row + 100):  # +100 to ensure we scan enough rows
            empty_count = 0
            empty_rows = []
            
            for check_row in range(row_num, row_num + 3):  # Check for 3 consecutive empty rows
                row_empty = True
                for col in range(2, 11):  # B to J columns
                    if ws.cell(row=check_row, column=col).value not in (None, ""):
                        row_empty = False
                        break
                
                if row_empty:
                    empty_count += 1
                    empty_rows.append(check_row)
                else:
                    break
            
            if empty_count == 3:
                # Use the second empty row
                current_row = empty_rows[1]
                logger.info(f"Found 3 consecutive empty rows, using row {current_row} for data entry")
                found = True
                break
        
        if not found:
            raise ValueError("Could not find 3 consecutive empty rows for data entry")
    
    # Add date to column A of the current row
    ws.cell(row=current_row, column=1).value = date
    
    # Now we have our current_row (either fer or a new empty row)
    # 1. Update the cell B with Bill Number or "BS"
    ws.cell(row=current_row, column=2).value = bill_no if bill_no else "BS"
    
    # 2. Update the cell C with Cheq No
    ws.cell(row=current_row, column=3).value = cheq_no
    
    # 3. Update the cell D with Acc No
    ws.cell(row=current_row, column=4).value = acc_no
    
    # 4. Update the cell E with employee name
    ws.cell(row=current_row, column=5).value = employee["name"]
    
    # 5. Update the cell E in previous row with institute
    ws.cell(row=current_row-1, column=5).value = institute
    
    # Always update cell F in current row with "Capital"
    ws.cell(row=current_row, column=6).value = "Capital"
    
    # Always update cell F in next row with "Interest"
    ws.cell(row=current_row+1, column=6).value = "Interest"
    
    # Handle capital amount if provided
    if capital_value is not None:
        # Determine which column to update based on bank name
        if bank_name == "HNB":
            cell = ws.cell(row=current_row, column=9)  # Column I
            cell.value = capital_value
            
        elif bank_name == "Peoples Bank":
            cell = ws.cell(row=current_row, column=8)  # Column H
            cell.value = capital_value
          
        elif bank_name == "Cash in Hand":
            cell = ws.cell(row=current_row, column=7)  # Column G
            cell.value = capital_value
    
    # Handle interest amount if provided
    if interest_value is not None:
        # Determine which column to update based on bank name
        if bank_name == "HNB":
            cell = ws.cell(row=current_row+1, column=9)  # Column I in next row
            cell.value = interest_value
           
        elif bank_name == "Peoples Bank":
            cell = ws.cell(row=current_row+1, column=8)  # Column H in next row
            cell.value = interest_value
            
        elif bank_name == "Cash in Hand":
            cell = ws.cell(row=current_row+1, column=7)  # Column G in next row
            cell.value = interest_value
    
    # 8. Update the cell M with Description
    if description:
        ws.cell(row=current_row, column=13).value = description  # Column M
    
    return current_row


@app.route('/submitPayment', methods=['POST'])
def submit_payment():
    try:
        data = request.json
        logger.info("Received payment data from frontend: %s", data)

        # Extract and validate required fields
        institute = data.get("institute")
        employee = data.get("employee")
        capital_amount = data.get("capitalAmount")
        interest_amount = data.get("interestAmount")
        cheq_no = data.get("cheqNo")
        acc_no = data.get("accNo")
        date = data.get("date")
        
        # Validate required fields
        if not all([institute, employee, cheq_no, acc_no, date]) or (capital_amount is None and interest_amount is None):
            return jsonify({"error": "Institution, Employee, Bill No, Cheq No, and Acc No are required. Either Capital or Interest amount must be provided."}), 400
        
        # Perform atomic Excel operation
        with atomic_excel_operation(EXCEL_FILE_PATH) as workbook:
            current_row = perform_payment_operation(workbook, data)
        
        # After successful Excel update, update personal account
        logger.info("Updating personal account for employee: %s of institution: %s", employee["name"], institute)
        personal_account_result = update_personal_account(
            employee_name=employee["name"],
            employee_accountNo=employee["accountNo"],
            institution_name=institute,
            date=date,
            capital=float(capital_amount) if capital_amount else None,
            interest=float(interest_amount) if interest_amount else None
        )

        logger.info("Updating the interest in trial balance for employee: %s of institution: %s", employee["name"], institute)
        update_trial_balance_interest_result = update_interest_trial_balance(
            employee_name=employee["name"],
            employee_accountNo=employee["accountNo"],
            institution_name=institute,
            date=date,
            capital=float(capital_amount) if capital_amount else None,
            interest=float(interest_amount) if interest_amount else None
        )

        logger.info("Updating the capital in trial balance for employee: %s of institution: %s", employee["name"], institute)
        update_trial_balance_capital_result = update_capital_trial_balance(
            employee_name=employee["name"],
            employee_accountNo=employee["accountNo"],
            institution_name=institute,
            date=date,
            capital=float(capital_amount) if capital_amount else None,
            interest=float(interest_amount) if interest_amount else None
        )

        if not personal_account_result["success"]:
            logger.error("Failed to update personal account: %s", personal_account_result["error"])
            # Note: We don't return here as the main Excel update was successful

        if not update_trial_balance_interest_result["success"]:
            logger.error("Failed to update trial balance interest %s", personal_account_result["error"])

        if not update_trial_balance_capital_result["success"]:
            logger.error("Failed to update trial balance capital: %s", personal_account_result["error"])

        # Return success message with the row that was updated
        return jsonify({
            "message": "Payment information updated successfully in Excel!",
            "row_updated": current_row,
            "personal_account_update": personal_account_result
        }), 200
        
    except Exception as e:
        logger.error("Error in submit_payment: %s", str(e))
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


# def perform_batch_payment_operation(workbook, data):
#     """
#     Separated batch payment logic to work with atomic operations
#     """
#     # Extract batch data
#     date = data.get("date")
#     first_entry = data.get("first_entry")
#     employees = data.get("employees", [])

#     # Convert first_entry to integer
#     fer = int(first_entry)
    
#     ws = workbook["Sheet1"]  # Using Sheet1 by default

#     # Track the current row for each iteration
#     current_row = fer
#     updated_rows = []

#     # Process each employee
#     for employee in employees:
#         # Extract employee data
#         institute = employee.get("institution")
#         name = employee.get("name")
#         capital_amount = employee.get("capitalAmount")
#         interest_amount = employee.get("interestAmount")
#         acc_no = employee.get("accNo")
#         bank_name = employee.get("bankName", "")
#         description = employee.get("description", "")

#         # Validate required fields for each employee
#         if not all([institute, name, acc_no]) or (capital_amount is None and interest_amount is None):
#             raise ValueError(f"Missing required fields for employee {name}")

#         # Convert amounts to float for numeric handling if they exist
#         capital_value = None
#         if capital_amount:
#             try:
#                 capital_value = float(capital_amount)
#             except ValueError:
#                 raise ValueError(f"Capital amount must be a valid number for employee {name}")
                
#         interest_value = None
#         if interest_amount:
#             try:
#                 interest_value = float(interest_amount)
#             except ValueError:
#                 raise ValueError(f"Interest amount must be a valid number for employee {name}")

#         # Check if the cells from Bfer to Jfer are empty
#         is_empty_row = True
#         for col in range(2, 11):  # B to J columns (2 to 10 in 0-based indexing)
#             cell_value = ws.cell(row=current_row, column=col).value
#             if cell_value not in (None, ""):
#                 is_empty_row = False
#                 break

#         # If the current row isn't empty, find three consecutive empty rows
#         if not is_empty_row:
#             logger.info(f"Row {current_row} is not empty, searching for three consecutive empty rows...")
#             found = False
            
#             for row_num in range(current_row, ws.max_row + 100):  # +100 to ensure we scan enough rows
#                 empty_count = 0
#                 empty_rows = []
                
#                 for check_row in range(row_num, row_num + 3):  # Check for 3 consecutive empty rows
#                     row_empty = True
#                     for col in range(2, 11):  # B to J columns
#                         if ws.cell(row=check_row, column=col).value not in (None, ""):
#                             row_empty = False
#                             break
                    
#                     if row_empty:
#                         empty_count += 1
#                         empty_rows.append(check_row)
#                     else:
#                         break
                
#                 if empty_count == 3:
#                     # Use the second empty row
#                     current_row = empty_rows[1]
#                     logger.info(f"Found 3 consecutive empty rows, using row {current_row} for data entry")
#                     found = True
#                     break
            
#             if not found:
#                 raise ValueError(f"Could not find 3 consecutive empty rows for employee {name}")

#         # Add date to column A of the current row
#         ws.cell(row=current_row, column=1).value = date

#         # Update the cells with employee data
#         ws.cell(row=current_row, column=2).value = "BS"  # Bill Number is always "BS"
#         ws.cell(row=current_row, column=3).value = ""    # Empty Cheque No
#         ws.cell(row=current_row, column=4).value = acc_no
#         ws.cell(row=current_row, column=5).value = name
#         ws.cell(row=current_row-1, column=5).value = institute

#         # Set payment types
#         ws.cell(row=current_row, column=6).value = "Capital"
#         ws.cell(row=current_row+1, column=6).value = "Interest"

#         # Handle capital amount if provided
#         if capital_value is not None:
#             if bank_name == "HNB":
#                 ws.cell(row=current_row, column=9).value = capital_value
#             elif bank_name == "Peoples Bank":
#                 ws.cell(row=current_row, column=8).value = capital_value
#             elif bank_name == "Cash in Hand":
#                 ws.cell(row=current_row, column=7).value = capital_value

#         # Handle interest amount if provided
#         if interest_value is not None:
#             if bank_name == "HNB":
#                 ws.cell(row=current_row+1, column=9).value = interest_value
#             elif bank_name == "Peoples Bank":
#                 ws.cell(row=current_row+1, column=8).value = interest_value
#             elif bank_name == "Cash in Hand":
#                 ws.cell(row=current_row+1, column=7).value = interest_value

#         # Add description if provided
#         if description:
#             ws.cell(row=current_row, column=13).value = description

#         # Track the updated row
#         updated_rows.append(current_row)

#         # Move to the next potential row (after the interest row)
#         current_row += 3

#     return updated_rows, employees


def perform_batch_payment_operation(workbook, data):
    """
    Enhanced batch payment logic with robust row availability checking
    """
    # Extract batch data
    date = data.get("date")
    first_entry = data.get("first_entry")
    employees = data.get("employees", [])

    # Convert first_entry to integer
    fer = int(first_entry)
    
    ws = workbook["Sheet1"]  # Using Sheet1 by default

    # Calculate required rows for the entire batch
    num_employees = len(employees)
    required_rows = num_employees * 3 + 3
    
    logger.info(f"Batch operation: {num_employees} employees, {required_rows} rows required")

    # Find the starting row (either fer if empty, or first available position)
    starting_row = fer
    
    # Check if the initial row (fer) is empty
    is_initial_row_empty = True
    for col in range(2, 11):  # B to J columns (2 to 10 in 0-based indexing)
        cell_value = ws.cell(row=starting_row, column=col).value
        if cell_value not in (None, ""):
            is_initial_row_empty = False
            break

    # If the initial row isn't empty, find three consecutive empty rows
    if not is_initial_row_empty:
        logger.info(f"Row {starting_row} is not empty, searching for three consecutive empty rows...")
        found = False
        
        for row_num in range(starting_row, ws.max_row + 100):  # +100 to ensure we scan enough rows
            empty_count = 0
            empty_rows = []
            
            for check_row in range(row_num, row_num + 3):  # Check for 3 consecutive empty rows
                row_empty = True
                for col in range(2, 11):  # B to J columns
                    if ws.cell(row=check_row, column=col).value not in (None, ""):
                        row_empty = False
                        break
                
                if row_empty:
                    empty_count += 1
                    empty_rows.append(check_row)
                else:
                    break
            
            if empty_count == 3:
                # Use the second empty row as starting point
                starting_row = empty_rows[1]
                logger.info(f"Found 3 consecutive empty rows, using row {starting_row} as starting point")
                found = True
                break
        
        if not found:
            raise ValueError("Could not find 3 consecutive empty rows to start the batch operation")

    # Now validate that we have enough consecutive empty rows for the entire batch
    logger.info(f"Validating {required_rows} consecutive empty rows starting from row {starting_row}")
    
    insufficient_rows = []
    for row_offset in range(required_rows):
        check_row = starting_row + row_offset
        row_empty = True
        
        for col in range(2, 11):  # B to J columns
            cell_value = ws.cell(row=check_row, column=col).value
            if cell_value not in (None, ""):
                row_empty = False
                insufficient_rows.append(check_row)
                break
    
    if insufficient_rows:
        error_message = (
            f"Insufficient empty rows for batch operation. "
            f"Required: {required_rows} consecutive empty rows starting from row {starting_row}. "
            f"Found non-empty data in rows: {insufficient_rows[:10]}"  # Limit to first 10 for readability
        )
        if len(insufficient_rows) > 10:
            error_message += f" and {len(insufficient_rows) - 10} more rows"
        
        logger.error(error_message)
        raise ValueError(error_message)

    # If we reach here, we have sufficient empty rows
    logger.info(f"Validation passed: {required_rows} consecutive empty rows available starting from row {starting_row}")

    # Track the current row for each iteration
    current_row = starting_row
    updated_rows = []

    # Process each employee
    for idx, employee in enumerate(employees):
        logger.info(f"Processing employee {idx + 1}/{num_employees}: {employee.get('name', 'Unknown')}")
        
        # Extract employee data
        institute = employee.get("institution")
        name = employee.get("name")
        capital_amount = employee.get("capitalAmount")
        interest_amount = employee.get("interestAmount")
        acc_no = employee.get("accNo")
        bank_name = employee.get("bankName", "")
        description = employee.get("description", "")

        # Validate required fields for each employee
        if not all([institute, name, acc_no]) or (capital_amount is None and interest_amount is None):
            raise ValueError(f"Missing required fields for employee {name}")

        # Convert amounts to float for numeric handling if they exist
        capital_value = None
        if capital_amount:
            try:
                capital_value = float(capital_amount)
            except ValueError:
                raise ValueError(f"Capital amount must be a valid number for employee {name}")
                
        interest_value = None
        if interest_amount:
            try:
                interest_value = float(interest_amount)
            except ValueError:
                raise ValueError(f"Interest amount must be a valid number for employee {name}")

        # Add date to column A of the current row
        ws.cell(row=current_row, column=1).value = date

        # Update the cells with employee data
        ws.cell(row=current_row, column=2).value = "BS"  # Bill Number is always "BS"
        ws.cell(row=current_row, column=3).value = ""    # Empty Cheque No
        ws.cell(row=current_row, column=4).value = acc_no
        ws.cell(row=current_row, column=5).value = name
        ws.cell(row=current_row-1, column=5).value = institute

        # Set payment types
        ws.cell(row=current_row, column=6).value = "Capital"
        ws.cell(row=current_row+1, column=6).value = "Interest"

        # Handle capital amount if provided
        if capital_value is not None:
            if bank_name == "HNB":
                ws.cell(row=current_row, column=9).value = capital_value
            elif bank_name == "Peoples Bank":
                ws.cell(row=current_row, column=8).value = capital_value
            elif bank_name == "Cash in Hand":
                ws.cell(row=current_row, column=7).value = capital_value

        # Handle interest amount if provided
        if interest_value is not None:
            if bank_name == "HNB":
                ws.cell(row=current_row+1, column=9).value = interest_value
            elif bank_name == "Peoples Bank":
                ws.cell(row=current_row+1, column=8).value = interest_value
            elif bank_name == "Cash in Hand":
                ws.cell(row=current_row+1, column=7).value = interest_value

        # Add description if provided
        if description:
            ws.cell(row=current_row, column=13).value = description

        # Track the updated row
        updated_rows.append(current_row)

        # Move to the next set of rows (each employee uses 3 rows)
        current_row += 3

    logger.info(f"Batch operation completed successfully. Updated rows: {updated_rows}")
    return updated_rows, employees



@app.route('/submitExcelBatchPayment', methods=['POST'])
def submit_batch_payment():
    try:
        data = request.json
        logger.info("Received batch payment data from frontend: %s", data)

        # Extract and validate batch data
        date = data.get("date")
        first_entry = data.get("first_entry")
        employees = data.get("employees", [])

        # Validate required fields
        if not all([date, first_entry, employees]):
            return jsonify({"error": "Date, first entry, and employees list are required"}), 400

        # Perform atomic Excel operation
        with atomic_excel_operation(EXCEL_FILE_PATH) as workbook:
            updated_rows, processed_employees = perform_batch_payment_operation(workbook, data)

        # After successful Excel update, update personal accounts
        personal_account_results = []
        for employee in processed_employees:
            logger.info("Updating personal account for employee: %s of institution: %s", 
                       employee.get("name"), employee.get("institution"))
            
            personal_account_result = update_personal_account(
                employee_name=employee.get("name"),
                employee_accountNo=employee.get("accNo"),
                institution_name=employee.get("institution"),
                date=date,
                capital=float(employee.get("capitalAmount")) if employee.get("capitalAmount") else None,
                interest=float(employee.get("interestAmount")) if employee.get("interestAmount") else None
            )

            logger.info("Updating trial balance interest for employee: %s of institution: %s", 
                       employee.get("name"), employee.get("institution"))
            update_trial_balance_interest_result = update_interest_trial_balance(
                employee_name=employee.get("name"),
                employee_accountNo=employee.get("accNo"),
                institution_name=employee.get("institution"),
                date=date,
                capital=float(employee.get("capitalAmount")) if employee.get("capitalAmount") else None,
                interest=float(employee.get("interestAmount")) if employee.get("interestAmount") else None
            )

            logger.info("Updating trial balance capital for employee: %s of institution: %s", 
                       employee.get("name"), employee.get("institution"))
            update_trial_balance_capital_result = update_capital_trial_balance(
                employee_name=employee.get("name"),
                employee_accountNo=employee.get("accNo"),
                institution_name=employee.get("institution"),
                date=date,
                capital=float(employee.get("capitalAmount")) if employee.get("capitalAmount") else None,
                interest=float(employee.get("interestAmount")) if employee.get("interestAmount") else None
            )

            if not personal_account_result["success"]:
                logger.error("Failed to update personal account for %s: %s", 
                           employee.get("name"), personal_account_result["error"])
            else:
                logger.info("Personal account update successful for %s: %s", 
                           employee.get("name"), personal_account_result["message"])

            if not update_trial_balance_interest_result["success"]:
                logger.error("Failed to update trial balance interest for %s: %s", 
                           employee.get("name"), update_trial_balance_interest_result["error"])
            else:
                logger.info("Trial balance interest update successful for %s: %s", 
                           employee.get("name"), update_trial_balance_interest_result["message"])

            if not update_trial_balance_capital_result["success"]:
                logger.error("Failed to update trial balance capital for %s: %s", 
                           employee.get("name"), update_trial_balance_capital_result["error"])
            else:
                logger.info("Trial balance capital update successful for %s: %s", 
                           employee.get("name"), update_trial_balance_capital_result["message"])

            personal_account_results.append({
                "employee": employee.get("name"),
                "result": personal_account_result
            })

        # Return success message with the rows that were updated
        return jsonify({
            "message": "Batch payment information updated successfully in Excel!",
            "rows_updated": updated_rows,
            "personal_account_updates": personal_account_results
        }), 200

    except Exception as e:
        logger.error("Error in submit_batch_payment: %s", str(e))
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':    
    app.run()